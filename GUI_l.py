import tkinter as tk
import openpyxl
import home


def login():
    def ok_click():
        wb = openpyxl.load_workbook("user_data.xlsx")
        ws = wb['Sheet1']
        name_col = "A"
        cal_col = "D"
        target_string = str(txt_u.get())
        for row in range(2, ws.max_row+1):
            if target_string in str(ws[f"{name_col}{row}"].value):
                a = str(ws[f"{name_col}{row}"].value)
                d = int(ws[f"{cal_col}{row}"].value)
                list = [a, d]

        with open('today_kcal.txt', 'w') as f:
            f.write(str(list[1]))
            f.write("\n")
        with open('user.txt', 'w') as f:
            f.write(str(list[0]))
            f.write("\n")
        home.home()
        root3.destroy()

    root3 = tk.Tk()

    root3.geometry('300x400')

    root3.title('ユーザー検索')

    lbl_n = tk.Label(root3, text='ユーザー名', foreground='Black')
    lbl_n.place(x=125, y=125)

    txt_u = tk.Entry(root3, width=25)
    txt_u.place(x=70, y=150)

    btn_u = tk.Button(root3, text='OK', command=ok_click)
    btn_u.place(x=125, y=200)

    root3.mainloop()
