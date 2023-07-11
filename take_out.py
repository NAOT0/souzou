import tkinter
import openpyxl
from pykakasi import kakasi
import pandas as pd
import yhoo
import math


def take_out():
    root = tkinter.Tk()
    root.geometry("320x240")
    root.title('商品取り出し')
    label = tkinter.Label(root, text='取り出す商品を読み取ってください', font=("", 10))
    label.pack()
    entry1 = tkinter.Entry(root)
    entry1.pack()
    label2 = tkinter.Label(root, text='何グラムか入力してください', font=("", 10))
    label2.pack()
    entry2 = tkinter.Entry(root)
    entry2.pack()
    button = tkinter.Button(root, text="OK")
    button.pack()

    def on_key1(event):
        entry2.focus_set()

    def on_key2(event):
        button.focus_set()
    entry1.bind("<Key-Return>", on_key1)
    entry2.bind("<Key-Return>", on_key2)

    def click():
        jancode_management = int(entry1.get())
        g = int(entry2.get())
        df = pd.read_excel('eiyouso2.xlsx', sheet_name=0)
        list = []
        B = []
        for row in df.values:
            list.append(f'{row[0]}')

        A = list
        kaki = kakasi()
        kaki.setMode('J', 'H')
        kaki.setMode("K", "H")
        conv = kaki.getConverter()
        sr = []
        sr.append(yhoo.reaaaad(jancode_management))

        X1 = sr[0]

        X1 = conv.do(X1)
        B.append(X1)

        Wb = openpyxl.load_workbook('eiyouso2.xlsx')
        Sheet_1 = Wb['シート1']
        Sheet_Max = Sheet_1.max_row
        data = []
        for i in range(len(A)):
            if A[i] in B[0]:
                data.append(Sheet_1["A"+str(i+2)].value)
                data.append(Sheet_1["B"+str(i+2)].value)

        Wb.save("eiyouso2.xlsx")
        all = data[1]
        g = g/100
        all_kcal = all*g
        all_kcal = math.floor(all_kcal)
        int(all_kcal)
        print(all_kcal)
        all_kcal = str(all_kcal)
        with open('ALL_kcal.txt', "a") as f:

            f.write(all_kcal + "\n")

        jajaja = jancode_management
        MOUIYADA = []
        MOUIYADA.append(jajaja)
        jancode_management = str(jancode_management)

        out_name_data = []
        Wb = openpyxl.load_workbook('jan_time.xlsx')
        Sheet_1 = Wb['Sheet1']
        Sheet_Max = Sheet_1.max_row
        for i in range(1, Sheet_Max):
            out_name_data.append(Sheet_1["A"+str(i+1)].value)
        for i in range(len(out_name_data)):
            if MOUIYADA[0] == out_name_data[i]:
                Sheet_1.delete_rows(i+2)

        Wb.save("jan_time.xlsx")
        entry1.delete(0, tkinter.END)
        entry2.delete(0, tkinter.END)
    button["command"] = click
    root.mainloop()
